# app_run_minimal_final.py
# 실행: streamlit run C:/Users/user/Desktop/스트림릿/app_run_minimal_final.py
import os
import random
from typing import List, Dict, Any
import streamlit as st
from openpyxl import load_workbook
# ----------------------
# 기본 설정
# ----------------------
st.set_page_config(page_title="MOMULAE - 강남역 맛집 추천", page_icon="🍱", layout="centered")
# 엑셀 경로(필요시 수정)
EXCEL_PATH = r"momulae_DB.xlsx"
IMG_DIR = os.path.dirname(EXCEL_PATH)  # 이미지 파일은 엑셀과 같은 폴더에 둡니다.
# ----------------------
# 데이터 로더 (openpyxl)
# 엑셀 구조: A:1차, B:2차, C:식당명, D:URL, E~: 키워드(선택)
# ----------------------
@st.cache_data(show_spinner=False)
def load_db(path: str) -> List[Dict[str, Any]]:
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    rows: List[Dict[str, Any]] = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r:
            continue
        c1 = "" if r[0] is None else str(r[0]).strip()
        c2 = "" if r[1] is None else str(r[1]).strip()
        nm = "" if r[2] is None else str(r[2]).strip()
        url = "" if r[3] is None else str(r[3]).strip()
        kws = []
        if len(r) > 4:
            for c in r[4:]:
                if c is not None and str(c).strip():
                    kws.append(str(c).strip())
        if c1 and c2 and nm:
            rows.append({"cat1": c1, "cat2": c2, "name": nm, "url": url, "keywords": kws})
    return rows
# ----------------------
# 유틸
# ----------------------
def get_cat1_list(db: List[Dict[str, Any]]) -> List[str]:
    return sorted({row["cat1"] for row in db if row["cat1"]})
def get_cat2_list(db: List[Dict[str, Any]], cat1: str) -> List[str]:
    return sorted({row["cat2"] for row in db if row["cat1"] == cat1 and row["cat2"]})
def pick_random(db: List[Dict[str, Any]], cat1: str=None, cat2: str=None) -> Dict[str, Any]:
    cand = [r for r in db if (not cat1 or r["cat1"]==cat1) and (not cat2 or r["cat2"]==cat2) and r["name"]]
    return random.choice(cand) if cand else {}
def tags_line(tags: List[str]) -> str:
    # 키워드 사이 구분자: 공백
    return " ".join([f"#{t.replace(' ','')}" for t in tags]) if tags else ""
def cat_image_path_exact(cat: str) -> str | None:
    """정확 일치 규칙: <카테고리명>.png 만 허용 (대소문자 포함 그대로)"""
    fname = f"{cat}.png"
    path = os.path.join(IMG_DIR, fname)
    return path if os.path.exists(path) else None
# ----------------------
# 세션 상태
# ----------------------
if "page" not in st.session_state: st.session_state.page = "home"
if "cat1" not in st.session_state: st.session_state.cat1 = None
if "cat2" not in st.session_state: st.session_state.cat2 = None
if "picked" not in st.session_state: st.session_state.picked = None
# ----------------------
# 콜백
# ----------------------
def cb_start(): st.session_state.page = "cat1"
def cb_back(to): st.session_state.page = to
def cb_reset():
    st.session_state.page = "home"
    st.session_state.cat1 = None
    st.session_state.cat2 = None
    st.session_state.picked = None
def cb_cat1(label: str, db: List[Dict[str, Any]]):
    # '랜덤' 선택 시 DB 전체에서 1곳 추천 → 바로 결과 화면
    if label == "랜덤":
        st.session_state.cat1 = "랜덤"
        st.session_state.cat2 = None
        st.session_state.picked = pick_random(db)  # 전체에서 랜덤
        st.session_state.page = "result"
        return
    # 기존 동작
    st.session_state.cat1 = label
    st.session_state.cat2 = None
    st.session_state.picked = None
    st.session_state.page = "cat2"
def cb_cat2(label: str, db: List[Dict[str, Any]]):
    st.session_state.cat2 = label
    st.session_state.picked = pick_random(db, st.session_state.cat1, label)
    st.session_state.page = "result"
# ----------------------
# 화면들 (중앙 배치: 가운데 열에만 UI)
# ----------------------
def page_home(db):
    st.container(height=80, border=False)
    _, mid, _ = st.columns([1,2,1])
    with mid:
        st.header("MOMULAE", anchor=False)
        st.write("")
        st.button("S T A R T", on_click=cb_start)
        st.caption("강남역 인근 맛집을 간단히 골라드려요.")
    st.container(height=60, border=False)
def page_cat1(db):
    _, mid, _ = st.columns([1,2,1])
    with mid:
        st.subheader("🍴 끌리는 음식 있어요?")
        cats = get_cat1_list(db)
        if not cats:
            st.warning("엑셀에 1차 카테고리가 없습니다.")
            st.button("처음부터", on_click=cb_reset)
            return
        # 1차 카테고리에 '랜덤' 항목 포함
        if "랜덤" not in cats:
            cats.append("랜덤")
        # 3열 그리드: 이미지(있으면) + 버튼(항상)
        for i in range(0, len(cats), 3):
            c1, c2, c3 = st.columns(3)
            row_items = cats[i:i+3]
            for j, label in enumerate(row_items):
                with (c1 if j==0 else c2 if j==1 else c3):
                    p = cat_image_path_exact(label)
                    if p: st.image(p, use_container_width=True)
                    st.button(label, use_container_width=True,
                              key=f"cat1_{label}", on_click=cb_cat1,
                              kwargs={"label": label, "db": db})
        st.button("이전", type="secondary", on_click=cb_back, kwargs={"to":"home"})
def page_cat2(db):
    """요청 반영: 안내 문구/캡션 수정 & '처음부터' 버튼 제거 유지"""
    _, mid, _ = st.columns([1,2,1])
    with mid:
        st.subheader("🍳 그럼 이중에서는요?")
        st.caption(f"{st.session_state.cat1 or ''}")
        subs = get_cat2_list(db, st.session_state.cat1 or "")
        if not subs:
            st.warning("해당 1차 카테고리에 2차 카테고리가 없습니다.")
        else:
            for i in range(0, len(subs), 3):
                c1, c2, c3 = st.columns(3)
                row_items = subs[i:i+3]
                for j, label in enumerate(row_items):
                    with (c1 if j==0 else c2 if j==1 else c3):
                        st.button(label, use_container_width=True,
                                  key=f"cat2_{label}", on_click=cb_cat2,
                                  kwargs={"label": label, "db": db})
        st.button("이전", on_click=cb_back, kwargs={"to":"cat1"})
def page_result(db):
    _, mid, _ = st.columns([1,2,1])
    with mid:
        cat1 = st.session_state.cat1 or ""
        cat2 = st.session_state.cat2 or ""
        st.subheader("🎁 그럼 여기 어때요?")
        st.caption(f"{cat1}  >  {cat2}" if cat2 else cat1)
        r = st.session_state.picked or {}
        if not r:
            st.error("추천할 식당이 없습니다. 카테고리를 다시 선택해 주세요.")
            st.button("처음부터", on_click=cb_reset)
            return
        name, url = r["name"], r["url"]
        # :흰색_확인_표시: 변경: 밑줄 복원 + 20px 유지 (링크 / 비링크 케이스 모두 처리)
        if url:
            st.markdown(
                f'<a href="{url}" target="_blank" '
                f'style="font-size:20px; font-weight:700; text-decoration:underline;">{name}</a>',
                unsafe_allow_html=True
            )
        else:
            st.markdown(
                f'<span style="font-size:20px; font-weight:700; text-decoration:underline;">{name}</span>',
                unsafe_allow_html=True
            )
        st.caption(tags_line(r.get("keywords", [])) or "등록된 키워드가 없어요.")
        c1, c2 = st.columns(2)
        with c1:
            if st.button("🔁 다시 추천"):
                # '랜덤' 상태면 전체에서 재추첨, 아니면 현재 필터 유지
                if (st.session_state.cat1 or "") == "랜덤":
                    st.session_state.picked = pick_random(db)
                else:
                    st.session_state.picked = pick_random(db, cat1, cat2)
        with c2:
            st.button("처음부터", on_click=cb_reset)
# ----------------------
# 엔트리
# ----------------------
def main():
    if not os.path.exists(EXCEL_PATH):
        st.error(f"엑셀 파일이 없습니다: {EXCEL_PATH}")
        return
    db = load_db(EXCEL_PATH)
    if not db:
        st.warning("엑셀에서 유효한 데이터가 없습니다.")
        return
    p = st.session_state.page
    if p == "home":   page_home(db)
    elif p == "cat1": page_cat1(db)
    elif p == "cat2": page_cat2(db)
    elif p == "result": page_result(db)
    else:             page_home(db)
if __name__ == "__main__":
    main()







